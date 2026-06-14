"""Build the catalog-curiosity spreadsheet (Sheets-importable XLSX).

Reframed around the actual cause: Movie Madness recycles WordPress rental posts,
leaving the previous title's URL slug and body text behind. So this lists the
records whose stored synopsis still describes the post's PREVIOUS occupant.
Source of truth = films.parquet's synopsis_overridden flag (the union detector
in stage 03). The URL slug is the evidence -- it names what the post used to be.
"""

import html
import re

import pandas as pd
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

SLUG_DROP = {"dvd", "dvdr", "vhs", "blu", "ray", "4k", "uhd", "criterion", "special",
             "edition", "collection", "set", "disc", "the", "a", "remastered", "uncut"}


def norm(s):
    return re.sub(r"[^a-z0-9]+", " ", str(s or "").lower()).strip()


def strip_html(s):
    return re.sub(r"\s+", " ", html.unescape(re.sub(r"<[^>]+>", " ", str(s or "")))).strip()


def prev_from_slug(slug):
    toks = [t for t in str(slug).split("-") if t and t not in SLUG_DROP and not (t.isdigit() and len(t) == 4)]
    return " ".join(w.capitalize() for w in toks) or str(slug)


F = pd.read_parquet("data/films.parquet")
flagged = F[F.synopsis_overridden == True].copy()  # noqa: E712  -- the 3,329 union set
raw = pd.read_parquet("data/catalog_raw.parquet")[["id", "slug", "title_raw", "content_html"]]
body = {r.id: norm(strip_html(r.content_html))[:200] for r in raw.itertuples()}
slug_by_id = dict(zip(raw.id, raw.slug))
title_by_id = dict(zip(raw.id, raw.title_raw))

rows = []
for r in flagged.itertuples():
    scn = norm(r.synopsis_catalog)
    skus = [s for s in (list(r.sku_ids) if r.sku_ids is not None else []) if body.get(s, "") == scn[:200]]
    if not skus:
        skus = list(r.sku_ids) if r.sku_ids is not None else []
    for sid in skus:
        sl = slug_by_id.get(sid, "")
        rows.append({
            "Fixed?": False,
            "WordPress post ID": int(sid),
            "Record title (now)": title_by_id.get(sid, r.title),
            "URL slug (what it used to be)": sl,
            "Previous occupant (from slug)": prev_from_slug(sl),
            "Stored synopsis (describes the OLD title)": str(r.synopsis_catalog),
            "Record actually is": f"{r.title} ({int(r.year) if pd.notna(r.year) else '?'})",
            "Suggested synopsis (TMDB)": str(r.overview),
            "Edit link (wp-admin)": f"https://www.moviemadness.org/wp-admin/post.php?post={sid}&action=edit",
        })

D = pd.DataFrame(rows).sort_values("Record title (now)")
print(f"records: {len(D)} (from {len(flagged)} films)")

about = pd.DataFrame({"Movie Madness — recycled rental records": [
    "What this is: rental records whose stored synopsis describes the title the post USED to be.",
    "",
    "Cause: the catalog reuses WordPress posts. When a title leaves the shelves, an old record is",
    "re-titled for the new one, but the original URL slug and body text are left behind. So a post",
    "now titled 'The Imitation Game' still lives at .../silicon-valley-season-6-dvd and still stores",
    "Silicon Valley's plot. The 'URL slug' column is the tell -- it names the previous occupant.",
    "",
    "This is NOT customer-visible: the store's search reads synopses from a separate system and shows",
    "the correct text. The staleness only lives in the raw post body (and old permalinks / the public",
    "REST API). So this is a curiosity / optional cleanup, not a customer-facing problem.",
    "",
    "Columns: post ID + edit link locate the record; 'Suggested synopsis' is the current title's entry",
    "from The Movie Database (TMDB) -- verify/rewrite as you like. Select 'Fixed?' and Insert > Checkbox",
    "to tick records off. Found by a community map project; not endorsed or certified by TMDB.",
]})

out = "analysis/movie_madness_recycled_records.xlsx"
with pd.ExcelWriter(out, engine="openpyxl") as xw:
    about.to_excel(xw, sheet_name="About", index=False)
    D.to_excel(xw, sheet_name="Recycled records", index=False)
    wb = xw.book
    fill = PatternFill("solid", fgColor="1F3B4D")
    font = Font(bold=True, color="FFFFFF")
    widths = {"Fixed?": 8, "WordPress post ID": 14, "Record title (now)": 30,
              "URL slug (what it used to be)": 34, "Previous occupant (from slug)": 28,
              "Stored synopsis (describes the OLD title)": 58, "Record actually is": 26,
              "Suggested synopsis (TMDB)": 58, "Edit link (wp-admin)": 40}
    for ws in wb.worksheets:
        ws.freeze_panes = "A2"
        if ws.title == "About":
            ws.column_dimensions["A"].width = 102
            continue
        for ci, cell in enumerate(ws[1], 1):
            cell.fill, cell.font = fill, font
            ws.column_dimensions[get_column_letter(ci)].width = widths.get(cell.value, 18)
        for rr in ws.iter_rows(min_row=2):
            for cell in rr:
                cell.alignment = Alignment(wrap_text=True, vertical="top")
print(f"wrote {out}")
